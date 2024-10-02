from flask import Flask, request, jsonify, send_from_directory
import os
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from flask_cors import CORS 

app = Flask(__name__)
CORS(app) 

# Configure upload folder
UPLOAD_FOLDER = './uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Function to clean and extract tables from PDF
def extract_clean_table(page):
    raw_table = page.extract_table()
    clean_table = []
    
    if raw_table:
        for row in raw_table:
            if row and any(cell.strip() if cell else '' for cell in row):
                clean_table.append([cell.strip() if cell else '' for cell in row])
                
    return clean_table

# Function to format Excel sheets
def format_excel_sheet(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)

    # Open the PDF and extract tables
    with pdfplumber.open(file_path) as pdf:
        for i, page in enumerate(pdf.pages):
            table = extract_clean_table(page)
            if table:
                df = pd.DataFrame(table[1:], columns=table[0])  # Assume first row as headers
                excel_file_path = os.path.join(app.config['UPLOAD_FOLDER'], f'table_{i}.xlsx')
                
                # Create a new workbook for each table
                wb = Workbook()
                ws = wb.active
                ws.title = f'Sheet{i}'

                # Write the DataFrame to the Excel sheet
                for r in df.itertuples(index=False):
                    ws.append(r)

                # Format the sheet
                format_excel_sheet(ws)

                # Save the Excel file
                wb.save(excel_file_path)

    return jsonify({'message': 'File processed successfully.'}), 200

@app.route('/uploads/<filename>', methods=['GET'])
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
