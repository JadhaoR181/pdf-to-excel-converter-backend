import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

# Path to your PDF file
pdf_path = 'SEM5.pdf'  # Adjust the path to your file

# Function to clean and extract tables more robustly
def extract_clean_table(page):
    raw_table = page.extract_table()
    clean_table = []
    
    if raw_table:
        # Handle multi-line and empty rows
        for row in raw_table:
            # Ensure row exists and contains at least one non-empty cell
            if row and any(cell.strip() if cell else '' for cell in row):  
                # Strip spaces and handle None cells
                clean_table.append([cell.strip() if cell else '' for cell in row])
                
    return clean_table

# Open the PDF and extract tables
with pdfplumber.open(r'D:/RGIT/SEM5.pdf') as pdf:
    tables = []
    for page in pdf.pages:
        table = extract_clean_table(page)
        if table:
            tables.append(table)

# Convert extracted tables to DataFrames
dfs = []
for table in tables:
    df = pd.DataFrame(table[1:], columns=table[0])  # Assume first row as headers
    dfs.append(df)

# Save the extracted tables to Excel
with pd.ExcelWriter('output_formatted.xlsx', engine='openpyxl') as writer:
    for i, df in enumerate(dfs):
        df.to_excel(writer, sheet_name=f'Sheet{i}', index=False)

# Load the workbook for formatting
wb = load_workbook('output_formatted.xlsx')

# Function to mimic the PDF formatting as much as possible
def format_excel_sheet(ws):
    # Set column widths based on content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2  # Adjust column width

    # Set headers to bold and center-aligned
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply thin borders and center-align the entire table
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Optionally, add background color to the header for better visibility
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Apply formatting to each sheet in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    format_excel_sheet(ws)

# Save the formatted workbook
wb.save('output_formatted.xlsx')

print("Excel sheet completed.")